"""
Extract the weekly cashflow data from the uploaded Excel file into a structured JSON.
"""
import json
import openpyxl
from datetime import datetime
from pathlib import Path

UPLOAD_DIR = Path(__file__).parent.parent / "uploads"
xlsx = next(UPLOAD_DIR.glob("*Weekly cashflow*"), None) or next(UPLOAD_DIR.glob("*weekly*"), None)
if not xlsx:
    raise FileNotFoundError("No weekly cashflow Excel found in uploads/")

print(f"Reading: {xlsx.name}")
wb = openpyxl.load_workbook(xlsx, data_only=True)
ws = wb["Weekly Cashflow"]

# ── Discover week-ending columns ────────────────────────────────────────────
# Row 3 has dates in cols B onward (col index 2...). Skip cols that are empty
# or not date-like. Some cols between B and the data start may be blank.
# Strategy: scan row 3 for date values.
DATE_ROW = 3
weeks = []       # list of (col_idx, date_str)
for col in ws.iter_cols(min_row=DATE_ROW, max_row=DATE_ROW, min_col=2):
    cell = col[0]
    if isinstance(cell.value, datetime):
        weeks.append((cell.column, cell.value.strftime("%Y-%m-%d")))
    elif isinstance(cell.value, (int, float)) and cell.value > 40000:
        # Excel serial date
        from openpyxl.utils.datetime import from_excel
        weeks.append((cell.column, from_excel(cell.value).strftime("%Y-%m-%d")))

print(f"Found {len(weeks)} weeks: {weeks[0][1]} → {weeks[-1][1]}")

week_cols = [c for c, _ in weeks]
week_dates = [d for _, d in weeks]

# ── Row definitions ──────────────────────────────────────────────────────────
# We define the rows we want to capture and their roles.
# Row number (1-based), label override (if different from cell), section/total/indent flags.
ROW_DEFS = [
    # (excel_row, id,                      label,                         kind,    indent)
    (4,  "opening_balance",   "Opening cash balance",           "total",   0),
    # Inflows
    (6,  "inflows_header",    "Inflows",                        "section", 0),
    (7,  "invoiced_header",   "Invoiced",                       "section", 1),
    (8,  "inv_character",     "Character",                      "leaf",    2),
    (9,  "inv_bi",            "BI",                             "leaf",    2),
    (10, "inv_bi2",           None,                             "leaf",    2),  # unnamed BI lines
    (11, "inv_bi3",           None,                             "leaf",    2),
    (12, "inv_bi4",           None,                             "leaf",    2),
    (13, "inv_bi5",           None,                             "leaf",    2),
    (14, "inv_bi6",           None,                             "leaf",    2),
    (15, "inv_bi7",           None,                             "leaf",    2),
    (16, "committed_header",  "Committed",                      "section", 1),
    (17, "com_cadre",         "Cadre Data Science Project",     "leaf",    2),
    (18, "com_character",     "Character Bio",                  "leaf",    2),
    (19, "com_ppd",           "PPD - Sanofi",                   "leaf",    2),
    (21, "pipeline_header",   "Not yet closed",                 "section", 1),
    (22, "pip_general",       "General pot",                    "leaf",    2),
    (23, "pip_bi",            "BI Internal project",            "leaf",    2),
    (24, "pip_character",     "Character - Glaucoma",           "leaf",    2),
    (25, "pip_regenxbio",     "Regenxbio",                      "leaf",    2),
    (26, "pip_vantage",       "Vantage Bioscience",             "leaf",    2),
    (28, "us_commission",     "US Salaries - commission",       "leaf",    1),
    (29, "ltd_deposit",       "Long term deposit cash",         "leaf",    1),
    (30, "rd_tax",            "R&D Tax Credit",                 "leaf",    1),
    (31, "invest_income",     "Investment income",              "leaf",    1),
    (32, "vat",               "VAT",                            "leaf",    1),
    (34, "total_inflow",      "Total inflow",                   "total",   0),
    # Outflows
    (36, "outflows_header",   "Outflows",                       "section", 0),
    (37, "altris_commission", "Altris Modeller Commission",     "leaf",    1),
    (38, "computers",         "Computers for Customers",        "leaf",    1),
    (40, "uk_salaries",       "UK Salaries",                    "leaf",    1),
    (41, "uk_pension",        "UK Pension",                     "leaf",    1),
    (42, "uk_salary_tax",     "UK Salary tax",                  "leaf",    1),
    (43, "uk_psa",            "PSA",                            "leaf",    1),
    (44, "uk_commission",     "UK Commission",                  "leaf",    1),
    (47, "us_salaries",       "US Salaries",                    "leaf",    1),
    (49, "us_cust_success",   "US Salaries - customer success", "leaf",    1),
    (50, "us_pension",        "US Pension",                     "leaf",    1),
    (51, "us_salary_tax",     "US Salary tax",                  "leaf",    1),
    (52, "us_medical",        "US Medical Insurance",           "leaf",    1),
    (53, "us_travel",         "US Travel",                      "leaf",    1),
    (54, "us_other",          "US Other",                       "leaf",    1),
    (56, "ga_header",         "G&A",                            "section", 1),
    (58, "ga_accountants",    "Accountants",                    "leaf",    2),
    (59, "ga_grenke",         "Grenke Leasing",                 "leaf",    2),
    (60, "ga_cf",             "CF Corporate Finance",           "leaf",    2),
    (61, "ga_fora",           "Fora Space",                     "leaf",    2),
    (62, "ga_insurance",      "Life Insurance / Health",        "leaf",    2),
    (63, "ga_bank",           "Bank fees",                      "leaf",    2),
    (64, "ga_baker_tilley",   "Baker Tilley",                   "leaf",    2),
    (65, "ga_it_hw",          "IT Software & Hardware",         "leaf",    2),
    (66, "ga_it_sec",         "IT Security & Compliance",       "leaf",    2),
    (67, "ga_ikites",         "Ikites - IT Services US",        "leaf",    2),
    (68, "ga_legal_edge",     "Legal Edge",                     "leaf",    2),
    (69, "ga_hr_lawyers",     "HR Lawyers",                     "leaf",    2),
    (70, "ga_ip_lawyers",     "IP Lawyers",                     "leaf",    2),
    (71, "ga_recruitment",    "Recruitment",                    "leaf",    2),
    (72, "ga_welfare",        "Staff Welfare",                  "leaf",    2),
    (73, "ga_travel",         "Travel",                         "leaf",    2),
    (74, "ga_bluejay",        "Bluejay Holdings",               "leaf",    2),
    (75, "ga_ppd_repay",      "PPD Repayment",                  "leaf",    2),
    (76, "ga_general",        "General Expenses",               "leaf",    2),
    (77, "ga_audit",          "Audit Fees",                     "leaf",    2),
    (78, "ga_insure",         "Insurance",                      "leaf",    2),
    (79, "ga_coaching",       "Coaching",                       "leaf",    2),
    (80, "ga_consultant",     "Consultant",                     "leaf",    2),
    (82, "sm_header",         "Sales & Marketing",              "section", 1),
    (83, "sm_advertising",    "Advertising",                    "leaf",    2),
    (84, "sm_exhibitions",    "Exhibitions",                    "leaf",    2),
    (85, "emr_nextgen",       "Nextgen EMR",                    "leaf",    2),
    (86, "emr_altris",        "Altris EMR",                     "leaf",    2),
    (87, "emr_dylan",         "Dylan Gilsenan EMR",             "leaf",    2),
    (88, "emr_eyemd",         "Eyemd EMR",                      "leaf",    2),
    (95, "fr_header",         "Fundraising",                    "section", 1),
    (96, "fr_sidecar",        "Sidecar Services",               "leaf",    2),
    (97, "fr_v2s",            "Venture to Success",             "leaf",    2),
    (98, "fr_alex",           "Alexander Champion",             "leaf",    2),
    (101,"total_outflows",    "Total Outflows",                 "total",   0),
    (103,"closing_balance",   "Closing cash balance",           "total",   0),
    (105,"closing_prior",     "Closing balance (prior forecast)","subtotal",0),
    # Bank accounts
    (110,"bank_header",       "Bank accounts (GBP)",            "section", 0),
    (111,"bank_us",           "US",                             "leaf",    1),
    (112,"bank_hsbc_usd",     "HSBC USD",                       "leaf",    1),
    (113,"bank_hsbc_gbp",     "HSBC GBP",                       "leaf",    1),
    (114,"bank_revolut_usd",  "Revolut USD",                    "leaf",    1),
    (117,"bank_telleroo",     "Telleroo",                       "leaf",    1),
    (118,"bank_revolut_gbp",  "Revolut GBP",                    "leaf",    1),
    (119,"bank_barclays",     "Barclays GBP",                   "leaf",    1),
    (120,"bank_total",        "Total cash (actual)",            "total",   0),
    (122,"bank_diff",         "Difference (model vs actual)",   "subtotal",0),
    (126,"insignis",          "Insignis (long-term deposit)",   "leaf",    1),
]

# ── Extract values ────────────────────────────────────────────────────────────
def cell_val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(v, 2)
    return None  # skip text values in data cells

rows_out = []
for excel_row, rid, label_override, kind, indent in ROW_DEFS:
    # Use label_override or read from col A
    if label_override:
        label = label_override
    else:
        label = str(ws.cell(row=excel_row, column=1).value or "").strip() or rid

    if kind == "section":
        rows_out.append({"id": rid, "label": label, "kind": kind, "indent": indent})
        continue

    values = [cell_val(ws, excel_row, col) for col in week_cols]
    # Only include the row if it has any non-None values
    has_data = any(v is not None for v in values)
    if not has_data and kind == "leaf":
        continue  # skip empty leaf rows

    rows_out.append({
        "id": rid,
        "label": label,
        "kind": kind,
        "indent": indent,
        "values": values,
    })

# ── Output ───────────────────────────────────────────────────────────────────
out = {
    "weeks": week_dates,
    "rows": rows_out,
    "currency": "GBP",
    "extractedFrom": xlsx.name,
}

out_path = Path(__file__).parent.parent / "src" / "weekly_cashflow_data.json"
with open(out_path, "w") as f:
    json.dump(out, f, indent=2)

print(f"Written {len(rows_out)} rows × {len(week_dates)} weeks → {out_path}")

# Quick sanity check
closing = next((r for r in rows_out if r["id"] == "closing_balance"), None)
if closing:
    vals = [v for v in closing["values"] if v is not None]
    if vals:
        print(f"Closing balance range: {min(vals):,.0f} → {max(vals):,.0f}")
        # Find first value and last 4 non-None
        print(f"First: {vals[0]:,.0f} | Last 4: {[f'{v:,.0f}' for v in vals[-4:]]}")
