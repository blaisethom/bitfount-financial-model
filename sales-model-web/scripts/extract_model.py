#!/usr/bin/env python3
"""Extract financial model data from Excel file."""

import openpyxl
from openpyxl.utils import get_column_letter

XLSX_PATH = "/home/node/workspace/bitfount-financial-model/sales-model-web/uploads/2026-06-11T08-13-43-354Z__v4-model.xlsx"

TARGET_SHEETS = [
    "Revenue Model auto",
    "P&L",
    "Employees",
    "Direct Costs",
    "G&A",
    "Engineering",
    "Sales and Marketing",
    "Product Support",
    "Budgets",
    "HubSpot Pivot",
]

def cell_value(cell):
    v = cell.value
    if v is None:
        return ""
    return str(v)

def row_is_empty(row):
    return all(cell.value is None for cell in row)

def dump_sheet(ws, max_rows=200):
    printed = 0
    for row in ws.iter_rows():
        if row_is_empty(row):
            continue
        vals = [cell_value(c) for c in row]
        # Strip trailing empty strings
        while vals and vals[-1] == "":
            vals.pop()
        if vals:
            print("\t".join(vals))
            printed += 1
        if printed >= max_rows:
            print(f"  ... (truncated at {max_rows} rows)")
            break

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
print("Available sheets:", wb.sheetnames)
print()

for sheet_name in TARGET_SHEETS:
    # Try exact match first, then case-insensitive
    ws = None
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        for sn in wb.sheetnames:
            if sn.strip().lower() == sheet_name.strip().lower():
                ws = wb[sn]
                sheet_name = sn
                break

    if ws is None:
        print(f"=== SHEET NOT FOUND: {sheet_name} ===")
        print(f"  Available: {wb.sheetnames}")
        print()
        continue

    print(f"{'='*80}")
    print(f"=== SHEET: {sheet_name} (dims: {ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}) ===")
    print(f"{'='*80}")
    dump_sheet(ws, max_rows=200)
    print()
